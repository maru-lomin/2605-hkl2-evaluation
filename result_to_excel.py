#!/usr/bin/env python3
"""Parse result .log files, group by doc_type, write {doc_type}.xlsx (kv + per-table sheets)."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any, DefaultDict, Dict, List, Tuple

from openpyxl import Workbook


def kv_value_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        if not value:
            return ""
        return ", ".join(str(v) for v in value)
    return str(value)


def prediction_from_log(path: Path) -> Tuple[Dict[str, Any], str]:
    http_prefix = "HTTP Response: "
    file_prefix = "File: "
    file_label: str | None = None
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.startswith(file_prefix):
                file_label = line[len(file_prefix) :].strip()
            elif line.startswith(http_prefix):
                payload = line[len(http_prefix) :].strip()
                data = json.loads(payload)
                pred = data.get("prediction") or {}
                label = file_label if file_label else path.name
                return pred, label
    raise ValueError(f"No '{http_prefix.strip()}' line in {path}")


def extract_doc_type(pred: Dict[str, Any]) -> str:
    for dt in pred.get("doc_types") or []:
        name = dt.get("doc_type")
        if name:
            return str(name)
    for page in pred.get("all_pages") or []:
        name = page.get("doc_type")
        if name:
            return str(name)
    return "unknown"


def safe_xlsx_filename_stem(doc_type: str) -> str:
    bad = '<>:"/\\|?*'
    s = doc_type.strip() or "unknown"
    for c in bad:
        s = s.replace(c, "_")
    if not s or s == ".":
        s = "unknown"
    return s


def collect_log_paths(paths: List[Path]) -> List[Path]:
    """Expand directories to their `*.log` files; accept `.log` files directly."""
    logs: List[Path] = []
    seen: set[Path] = set()
    for raw in paths:
        p = raw.expanduser()
        if not p.exists():
            raise FileNotFoundError(f"Path not found: {p}")
        if p.is_dir():
            for f in sorted(p.glob("*.log")):
                key = f.resolve()
                if key not in seen:
                    seen.add(key)
                    logs.append(f)
        elif p.is_file():
            if p.suffix.lower() != ".log":
                raise ValueError(f"Expected a directory or a .log file, got: {p}")
            key = p.resolve()
            if key not in seen:
                seen.add(key)
                logs.append(p)
        else:
            raise ValueError(f"Not a file or directory: {p}")
    if not logs:
        raise ValueError(
            "No .log files found. Pass a directory containing *.log or .log file paths."
        )
    return sorted(logs, key=lambda x: str(x.resolve()))


def collect_kv_keys(predictions: List[Dict[str, Any]]) -> List[str]:
    keys: List[str] = []
    seen = set()
    for pred in predictions:
        for item in pred.get("key_values") or []:
            k = item.get("key")
            if k is None or k in seen:
                continue
            seen.add(k)
            keys.append(k)
    return keys


def row_kv_values(pred: Dict[str, Any], kv_keys: List[str]) -> List[str]:
    by_key = {item.get("key"): item for item in (pred.get("key_values") or [])}
    return [kv_value_to_str(by_key.get(k, {}).get("value")) for k in kv_keys]


def sanitize_sheet_name(name: str) -> str:
    bad_chars = ["\\", "/", "*", "?", ":", "[", "]"]
    for ch in bad_chars:
        name = name.replace(ch, "_")
    return name[:31] if len(name) > 31 else name


def infer_key_codes(table: Dict[str, Any]) -> List[str]:
    ch = table.get("column-header") or {}
    key_codes = list(ch.get("key_code") or [])
    body = (table.get("body") or {}).get("content") or []
    if key_codes:
        return key_codes
    if body and body[0]:
        return [f"column_{i + 1}" for i in range(len(body[0]))]
    return []


def column_lists_for_table(table: Dict[str, Any]) -> Tuple[List[str], Dict[str, List[str]]]:
    key_codes = infer_key_codes(table)
    body = (table.get("body") or {}).get("content") or []
    lists: Dict[str, List[str]] = {kc: [] for kc in key_codes}
    for row in body:
        for i, kc in enumerate(key_codes):
            cell = row[i] if i < len(row) else ""
            lists[kc].append("" if cell is None else str(cell))
    return key_codes, lists


def merge_column_lists(
    acc: Dict[str, List[str]], key_codes: List[str], lists: Dict[str, List[str]]
) -> None:
    for kc in key_codes:
        acc.setdefault(kc, [])
        acc[kc].extend(lists.get(kc, []))


def iter_tables_by_key(pred: Dict[str, Any], table_key: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for page in pred.get("all_pages") or []:
        for table in page.get("tables") or []:
            if table.get("key") == table_key:
                out.append(table)
    return out


def collect_all_table_keys(predictions: List[Dict[str, Any]]) -> List[str]:
    keys: List[str] = []
    seen = set()
    for pred in predictions:
        for page in pred.get("all_pages") or []:
            for table in page.get("tables") or []:
                k = table.get("key")
                if k is None or k in seen:
                    continue
                seen.add(k)
                keys.append(k)
    return keys


def ordered_key_codes_for_table(table_key: str, predictions: List[Dict[str, Any]]) -> List[str]:
    ordered: List[str] = []
    seen = set()
    for pred in predictions:
        for table in iter_tables_by_key(pred, table_key):
            for kc in infer_key_codes(table):
                if kc not in seen:
                    seen.add(kc)
                    ordered.append(kc)
    return ordered


def build_table_row(
    pred: Dict[str, Any], table_key: str, column_keys: List[str]
) -> List[str]:
    merged: Dict[str, List[str]] = {}
    for table in iter_tables_by_key(pred, table_key):
        kc, lists = column_lists_for_table(table)
        merge_column_lists(merged, kc, lists)
    return [json.dumps(merged.get(k, []), ensure_ascii=False) for k in column_keys]


def unique_sheet_title(base: str, used: set) -> str:
    name = sanitize_sheet_name(base)
    if not name:
        name = "sheet"
    candidate = name
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = sanitize_sheet_name(name[: 31 - len(suffix)] + suffix)
        n += 1
    used.add(candidate)
    return candidate


def write_workbook(
    file_labels: List[str], predictions: List[Dict[str, Any]], output_path: Path
) -> None:
    kv_keys = collect_kv_keys(predictions)
    table_keys = collect_all_table_keys(predictions)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_names: set = set()
    kv_title = unique_sheet_title("kv", used_names)
    ws_kv = wb.create_sheet(kv_title)
    ws_kv.append(["파일명"] + kv_keys)
    for label, pred in zip(file_labels, predictions):
        ws_kv.append([label] + row_kv_values(pred, kv_keys))

    for table_key in table_keys:
        column_keys = ordered_key_codes_for_table(table_key, predictions)
        if not column_keys:
            column_keys = ["column_1"]
        sheet_title = unique_sheet_title(table_key, used_names)
        ws = wb.create_sheet(sheet_title)
        ws.append(["파일명"] + column_keys)
        for label, pred in zip(file_labels, predictions):
            ws.append([label] + build_table_row(pred, table_key, column_keys))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def convert_logs_to_doc_type_workbooks(
    log_paths: List[Path], out_dir: Path
) -> List[Path]:
    by_doc: DefaultDict[str, List[Tuple[str, Dict[str, Any]]]] = defaultdict(list)
    for path in log_paths:
        try:
            pred, file_label = prediction_from_log(path)
        except ValueError as exc:
            # Skip malformed logs and continue processing the rest.
            print(f"Skipped: {path} ({exc})")
            continue
        doc_type = extract_doc_type(pred)
        by_doc[doc_type].append((file_label, pred))

    written: List[Path] = []
    for doc_type, rows in sorted(by_doc.items()):
        labels = [r[0] for r in rows]
        preds = [r[1] for r in rows]
        stem = safe_xlsx_filename_stem(doc_type)
        out_path = out_dir / f"{stem}.xlsx"
        write_workbook(labels, preds, out_path)
        written.append(out_path)
    return written


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse result .log files, group by doc_type, write {doc_type}.xlsx per group."
    )
    parser.add_argument(
        "input_path",
        nargs="+",
        help="One or more directories (all *.log inside) or .log file paths",
    )
    parser.add_argument(
        "-d",
        "--out-dir",
        type=Path,
        default=Path("."),
        help="Directory for output workbooks (default: current directory)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    paths = collect_log_paths([Path(p) for p in args.input_path])
    written = convert_logs_to_doc_type_workbooks(paths, args.out_dir)
    for w in written:
        print(f"Saved: {w}")


if __name__ == "__main__":
    main()
