#!/usr/bin/env python3
"""Map kv_class, doc_type, and table sheets (table_map.json) in xlsx files."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

KV_CLASS_MAP_FILE = "kv_class_map.json"
DOC_TYPE_MAP_FILE = "doc_type_map.json"
DOC_TYPE_TO_KV_FILE = "doc_type_to_kv.json"
TABLE_MAP_FILE = "table_map.json"
HEADER_PREFIXES = ("pred-", "gt-", "tp-", "fp-", "fn-")


def safe_filename(name: str) -> str:
    bad = '<>:"/\\|?*'
    out = name.strip() or "unknown"
    for ch in bad:
        out = out.replace(ch, "_")
    return out or "unknown"


def load_mapping(path: Path) -> dict[str, str]:
    if not path.is_file():
        raise FileNotFoundError(f"Mapping file not found: {path}")
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise ValueError(f"{path} must contain a JSON object (string → string).")
    return {str(k).strip(): str(v) for k, v in data.items() if k is not None}


def load_table_map(path: Path) -> dict[str, dict[str, Any]]:
    """Table key (sheet name) → { sheet: 한글 시트명, kv: { 항목코드: 필드명 } }. Missing file → {}."""
    if not path.is_file():
        return {}
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise ValueError(f"{path} must contain a JSON object.")
    out: dict[str, dict[str, Any]] = {}
    for k, v in data.items():
        key = str(k).strip()
        if not key or not isinstance(v, dict):
            continue
        kv_raw = v.get("kv")
        kv_inner: dict[str, str] = {}
        if isinstance(kv_raw, dict):
            for kk, vv in kv_raw.items():
                if kk is None:
                    continue
                kv_inner[str(kk).strip()] = "" if vv is None else str(vv)
        sheet_label = v.get("sheet")
        out[key] = {
            "sheet": "" if sheet_label is None else str(sheet_label).strip(),
            "kv": kv_inner,
        }
    return out


def merged_kv_map_for_sheet(
    sheet_title: str, kv_map: dict[str, str], table_map: dict[str, dict[str, Any]]
) -> dict[str, str]:
    """kv_class_map + table_map.kv for this sheet when the sheet is a known table."""
    entry = table_map.get(sheet_title)
    if not entry or not entry.get("kv"):
        return kv_map
    m = dict(kv_map)
    m.update(entry["kv"])
    return m


def trailing_number_sort_key(header: Any) -> tuple:
    """Sort key: ascending by the integer at the end of the header (e.g. …-001 → 1)."""
    s = "" if header is None else str(header).strip()
    base = s
    for prefix in HEADER_PREFIXES:
        if s.startswith(prefix):
            base = s[len(prefix) :].strip()
            break
    m = re.search(r"(\d+)\s*$", base)
    if m:
        return (0, int(m.group(1)), s)
    return (1, s)


def is_plain_result_kv_sheet_row1(headers: list[Any], kv_map: dict[str, str]) -> bool:
    """True for result_to_excel-style kv sheet: `파일명` + bare 항목코드 columns only.

    Excludes evaluation workbooks where row1 is `파일명`, `total`, `pred-*`, `gt-*`, … — reordering
    those by trailing number shuffles metric columns and misaligns cell values with headers.
    """
    if len(headers) < 2 or str(headers[0] or "").strip() != "파일명":
        return False
    seen_code = False
    for h in headers[1:]:
        if h is None:
            continue
        s = str(h).strip()
        if not s:
            continue
        if s == "total":
            return False
        if any(s.startswith(p) for p in HEADER_PREFIXES):
            return False
        if s not in kv_map:
            return False
        seen_code = True
    return seen_code


def map_kv_header_cell(raw: str, m: dict[str, str]) -> str:
    s = raw.strip()
    if not s:
        return raw
    for prefix in HEADER_PREFIXES:
        if s.startswith(prefix):
            code = s[len(prefix) :].strip()
            if code in m:
                return prefix + m[code]
            return raw
    if s in m:
        return m[s]
    return raw


def column_index_for_header(ws, name: str) -> int | None:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row1:
        return None
    for j, val in enumerate(row1):
        if val is None:
            continue
        if str(val).strip() == name:
            return j + 1
    return None


def unique_sheet_title(base: str, used: set[str]) -> str:
    name = base[:31] if len(base) > 31 else base
    if not name:
        name = "sheet"
    candidate = name
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = (name[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(candidate)
    return candidate


def expected_kv_codes_for_doc_type(
    문서명: str,
    doc_type_to_kv: dict[str, str],
    kv_map: dict[str, str],
) -> list[str] | None:
    """Return sorted kv_class codes in kv_map whose key starts with the prefix for 문서명."""
    prefix = doc_type_to_kv.get(문서명)
    if prefix is None:
        return None
    prefix = str(prefix).strip()
    if not prefix:
        return None
    codes = [k for k in kv_map if str(k).startswith(prefix)]
    codes.sort(key=trailing_number_sort_key)
    return codes


def augment_kv_sheet_missing_columns(
    wb,
    file_stem: str,
    kv_map: dict[str, str],
    doc_map: dict[str, str],
    doc_type_to_kv: dict[str, str],
) -> list[str]:
    """Append missing expected kv columns on sheet `kv`. Returns list of added codes (possibly empty)."""
    문서명 = doc_map.get(file_stem)
    if not 문서명:
        return []
    expected = expected_kv_codes_for_doc_type(문서명, doc_type_to_kv, kv_map)
    if not expected:
        return []
    if "kv" not in wb.sheetnames:
        return []
    ws = wb["kv"]
    ncols = ws.max_column or 0
    if ncols < 1:
        return []
    row1 = [ws.cell(1, c).value for c in range(1, ncols + 1)]
    if not row1 or str(row1[0] or "").strip() != "파일명":
        return []
    if not is_plain_result_kv_sheet_row1(row1, kv_map):
        return []
    existing: set[str] = set()
    for h in row1[1:]:
        if h is None:
            continue
        existing.add(str(h).strip())
    missing = [c for c in expected if c not in existing]
    if not missing:
        return []
    start_col = ws.max_column + 1
    for i, code in enumerate(missing):
        col_idx = start_col + i
        ws.cell(row=1, column=col_idx, value=code)
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx, value="")
    return missing


def reorder_kv_sheet_columns_by_trailing_number(ws, kv_map: dict[str, str]) -> None:
    """Reorder columns after `파일명` on sheet `kv` by trailing number in kv_class (ascending)."""
    if ws.max_row < 1 or ws.max_column < 2:
        return
    ncols = ws.max_column
    headers = [ws.cell(1, c).value for c in range(1, ncols + 1)]
    if not is_plain_result_kv_sheet_row1(headers, kv_map):
        return
    rest = headers[1:]
    if not rest:
        return
    order = list(range(len(rest)))
    order.sort(key=lambda i: trailing_number_sort_key(rest[i]))
    if order == list(range(len(rest))):
        return
    nrows = ws.max_row
    grid: list[list[Any]] = []
    for r in range(1, nrows + 1):
        grid.append([ws.cell(r, c).value for c in range(1, ncols + 1)])
    for r in range(nrows):
        tail = grid[r][1:]
        new_tail = [tail[i] for i in order]
        grid[r] = [grid[r][0]] + new_tail
    # openpyxl: ws.cell(..., value=None) does not clear the cell; assign .value so
    # reordered columns that should be empty do not keep stale text from old positions.
    for r in range(nrows):
        for c in range(len(grid[r])):
            ws.cell(row=r + 1, column=c + 1).value = grid[r][c]


def map_confusion_matrix_sheet(ws, doc_map: dict[str, str]) -> int:
    """Map first row (gt\\pred row) and first column labels if they are doc type codes."""
    n = 0
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for c in range(2, max_col + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is None:
            continue
        key = str(cell.value).strip()
        if key in doc_map:
            cell.value = doc_map[key]
            n += 1
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=1)
        if cell.value is None:
            continue
        key = str(cell.value).strip()
        if key in doc_map:
            cell.value = doc_map[key]
            n += 1
    return n


def apply_mapping_to_workbook(
    path: Path,
    kv_map: dict[str, str],
    doc_map: dict[str, str],
    doc_type_to_kv: dict[str, str],
    table_map: dict[str, dict[str, Any]],
    file_stem: str,
) -> tuple[int, int, int, int, int, list[str]]:
    """Returns (kv_headers, kv_class_cells, doc_type_cells, cm_cells, sheets_renamed, added_kv_cols)."""
    wb = load_workbook(path)
    added_kv_cols = augment_kv_sheet_missing_columns(
        wb, file_stem, kv_map, doc_map, doc_type_to_kv
    )
    if "kv" in wb.sheetnames:
        reorder_kv_sheet_columns_by_trailing_number(wb["kv"], kv_map)
    kv_headers_n = 0
    kv_cells_n = 0
    doc_cells_n = 0
    cm_cells_n = 0
    sheets_renamed = 0

    for ws in wb.worksheets:
        m = merged_kv_map_for_sheet(ws.title, kv_map, table_map)
        for cell in ws[1]:
            if cell.value is None:
                continue
            old = cell.value
            if not isinstance(old, str):
                old = str(old)
            new = map_kv_header_cell(old, m)
            if new != old:
                cell.value = new
                kv_headers_n += 1

        col_kv = column_index_for_header(ws, "kv_class")
        if col_kv is not None:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_kv)
                if cell.value is None:
                    continue
                key = str(cell.value).strip()
                if key in m:
                    cell.value = m[key]
                    kv_cells_n += 1

        col_dt = column_index_for_header(ws, "doc_type")
        if col_dt is not None:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_dt)
                if cell.value is None:
                    continue
                key = str(cell.value).strip()
                if key in doc_map:
                    cell.value = doc_map[key]
                    doc_cells_n += 1

        if ws.title == "confusion_matrix":
            cm_cells_n += map_confusion_matrix_sheet(ws, doc_map)

    used_names = set(wb.sheetnames)
    for ws in list(wb.worksheets):
        title = ws.title
        if title not in table_map:
            continue
        entry = table_map[title]
        new_label = entry.get("sheet") or ""
        if not new_label:
            continue
        used_names.discard(title)
        new_title = unique_sheet_title(safe_filename(new_label), used_names)
        if new_title != title:
            ws.title = new_title
            sheets_renamed += 1

    used_names = set(wb.sheetnames)
    for ws in list(wb.worksheets):
        title = ws.title
        if title not in doc_map:
            continue
        new_base = safe_filename(doc_map[title])
        used_names.discard(title)
        new_title = unique_sheet_title(new_base, used_names)
        if new_title != title:
            ws.title = new_title
            sheets_renamed += 1

    wb.save(path)
    wb.close()
    return kv_headers_n, kv_cells_n, doc_cells_n, cm_cells_n, sheets_renamed, added_kv_cols


def collect_xlsx_files(root: Path) -> list[Path]:
    """Only `root/*.xlsx` (depth 1)."""
    return sorted(root.glob("*.xlsx"))


def reserve_output_name(
    src: Path, doc_map: dict[str, str], used_names: set[str]
) -> str:
    """Target filename: stem mapped to 문서명 when known; resolve collisions."""
    stem = src.stem
    if stem in doc_map:
        base = safe_filename(doc_map[stem])
    else:
        base = safe_filename(stem)
    candidate = f"{base}.xlsx"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate
    n = 2
    while True:
        alt = f"{base}_{n}.xlsx"
        if alt not in used_names:
            used_names.add(alt)
            return alt
        n += 1


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Apply kv_class_map.json, doc_type_map.json, doc_type_to_kv.json, and table_map.json "
            "(beside this script) to INPUT_DIR/*.xlsx only. Output: sibling {INPUT_DIR.name}_mapped/. "
            "Table sheets listed in table_map.json get sheet names and column headers from that file. "
            "Missing kv columns on `kv` use doc_type_to_kv prefixes; columns are ordered by trailing "
            "number in each kv_class code, then mapping is applied."
        )
    )
    parser.add_argument(
        "input_dir",
        type=Path,
        help="Directory containing .xlsx files (only top-level *.xlsx)",
    )
    args = parser.parse_args()
    input_dir = args.input_dir.expanduser().resolve()
    if not input_dir.is_dir():
        print(f"Not a directory: {input_dir}", file=sys.stderr)
        sys.exit(1)

    script_dir = Path(__file__).resolve().parent
    try:
        kv_map = load_mapping(script_dir / KV_CLASS_MAP_FILE)
        doc_map = load_mapping(script_dir / DOC_TYPE_MAP_FILE)
        doc_type_to_kv = load_mapping(script_dir / DOC_TYPE_TO_KV_FILE)
        table_map = load_table_map(script_dir / TABLE_MAP_FILE)
    except (OSError, ValueError) as e:
        print(e, file=sys.stderr)
        sys.exit(1)

    files = collect_xlsx_files(input_dir)
    if not files:
        print(f"No .xlsx files in {input_dir} (expecting {input_dir}/*.xlsx)", file=sys.stderr)
        sys.exit(0)

    out_root = input_dir.parent / f"{input_dir.name}_mapped"
    out_root.mkdir(parents=True, exist_ok=True)

    used_out_names: set[str] = set()
    for src in files:
        out_name = reserve_output_name(src, doc_map, used_out_names)
        dst = out_root / out_name
        dst.write_bytes(src.read_bytes())
        kh, kk, dd, cm, sr, added = apply_mapping_to_workbook(
            dst, kv_map, doc_map, doc_type_to_kv, table_map, src.stem
        )
        if added:
            dn = doc_map.get(src.stem, src.stem)
            print(
                f"[missing kv columns] doc_type={dn} (stem={src.stem}): "
                f"added {len(added)} column(s)"
            )
        print(
            f"OK {src.name} -> {out_name}  "
            f"(kv_headers={kh}, kv_class={kk}, doc_type={dd}, "
            f"confusion_matrix={cm}, sheets_renamed={sr}, kv_cols_added={len(added)})"
        )

    print(f"Done. Output root: {out_root}")


if __name__ == "__main__":
    main()
