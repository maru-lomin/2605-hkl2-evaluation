#!/usr/bin/env python3
"""Evaluate pred vs gt XLSX files and write per-doc_type Excel reports."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

import matplotlib.pyplot as plt
from matplotlib import font_manager
from openpyxl import Workbook, load_workbook

_MATPLOTLIB_FONT_READY = False


def ensure_matplotlib_korean_font() -> None:
    """Pick a font with Hangul glyphs (wheel-bundled or system) for confusion matrix labels."""
    global _MATPLOTLIB_FONT_READY
    if _MATPLOTLIB_FONT_READY:
        return

    # PyPI wheel ships Nanum Gothic — works offline after `uv sync` with no OS fonts package.
    try:
        import koreanize_matplotlib  # noqa: F401
    except ImportError:
        pass
    else:
        _MATPLOTLIB_FONT_READY = True
        return

    font_entries: List[Tuple[str, List[str]]] = [
        (
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            ["Noto Sans CJK KR", "Noto Sans CJK JP", "Noto Sans CJK SC"],
        ),
        (
            "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
            ["NanumGothic", "Nanum Gothic"],
        ),
    ]

    for font_path, families in font_entries:
        path = Path(font_path)
        if not path.is_file():
            continue
        try:
            font_manager.fontManager.addfont(str(path))
        except OSError:
            continue
        seen: set[str] = set()
        merged: List[str] = []
        for name in families + list(plt.rcParams.get("font.sans-serif", ())):
            if name not in seen:
                merged.append(name)
                seen.add(name)
        plt.rcParams["font.sans-serif"] = merged
        plt.rcParams["font.family"] = "sans-serif"
        plt.rcParams["axes.unicode_minus"] = False
        _MATPLOTLIB_FONT_READY = True
        return

    for family in ("Noto Sans CJK KR", "NanumGothic", "Noto Sans CJK JP"):
        path = font_manager.findfont(font_manager.FontProperties(family=family))
        if path and "dejavu" not in path.lower():
            plt.rcParams["font.sans-serif"] = [family] + list(
                plt.rcParams.get("font.sans-serif", ())
            )
            plt.rcParams["font.family"] = "sans-serif"
            plt.rcParams["axes.unicode_minus"] = False
            _MATPLOTLIB_FONT_READY = True
            return

    _MATPLOTLIB_FONT_READY = True


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate OCR extraction with syllable-level micro F1."
    )
    parser.add_argument(
        "pred_dir",
        type=Path,
        nargs="?",
        default=Path("result"),
        help="Prediction XLSX directory (default: result)",
    )
    parser.add_argument(
        "gt_dir",
        type=Path,
        nargs="?",
        default=Path("ground_truth"),
        help="Ground-truth XLSX directory (default: ground_truth)",
    )
    parser.add_argument(
        "eval_dir",
        type=Path,
        nargs="?",
        default=Path("evaluation"),
        help="Output evaluation directory (default: evaluation)",
    )
    return parser.parse_args()


def xlsx_files_by_stem(directory: Path) -> Dict[str, Path]:
    files = {p.stem: p for p in sorted(directory.glob("*.xlsx"))}
    return files


def safe_filename(name: str) -> str:
    bad = '<>:"/\\|?*'
    out = name.strip() or "unknown"
    for ch in bad:
        out = out.replace(ch, "_")
    return out or "unknown"


def to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def parse_list_cell(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [to_string(v) for v in value]
    s = str(value).strip()
    if not s:
        return []
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [to_string(v) for v in parsed]
        except Exception:
            pass
    return [s]


def read_kv_sheet(workbook_path: Path) -> Dict[str, Dict[str, str]]:
    wb = load_workbook(workbook_path, data_only=True)
    if "kv" not in wb.sheetnames:
        return {}
    ws = wb["kv"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [to_string(v) for v in rows[0]]
    kv_keys = header[1:]
    result: Dict[str, Dict[str, str]] = {}
    for row in rows[1:]:
        if not row:
            continue
        file_name = to_string(row[0])
        if not file_name:
            continue
        values = list(row[1:])
        kv_data: Dict[str, str] = {}
        for i, key in enumerate(kv_keys):
            cell_value = values[i] if i < len(values) else None
            kv_data[key] = to_string(cell_value)
        result[file_name] = kv_data
    return result


def kv_classes_in_sheet(workbook_path: Path) -> List[str]:
    wb = load_workbook(workbook_path, data_only=True)
    if "kv" not in wb.sheetnames:
        return []
    ws = wb["kv"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [to_string(v).strip() for v in rows[0]]
    return [h for h in headers[1:] if h]


def read_table_sheets(workbook_path: Path) -> Dict[str, Dict[str, Dict[str, List[str]]]]:
    wb = load_workbook(workbook_path, data_only=True)
    table_data: Dict[str, Dict[str, Dict[str, List[str]]]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in {"kv", "summary"}:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [to_string(v) for v in rows[0]]
        class_keys = headers[1:]
        by_file: Dict[str, Dict[str, List[str]]] = {}
        for row in rows[1:]:
            if not row:
                continue
            file_name = to_string(row[0])
            if not file_name:
                continue
            values = list(row[1:])
            class_map: Dict[str, List[str]] = {}
            for i, class_key in enumerate(class_keys):
                cell_value = values[i] if i < len(values) else None
                class_map[class_key] = parse_list_cell(cell_value)
            by_file[file_name] = class_map
        table_data[sheet_name] = by_file
    return table_data


def char_counter(text: str) -> Counter:
    return Counter(list(text))


def diff_strings(pred_text: str, gt_text: str) -> Tuple[str, str, str, int, int, int]:
    pred_counter = char_counter(pred_text)
    gt_counter = char_counter(gt_text)
    keys = set(pred_counter) | set(gt_counter)

    tp_chars: List[str] = []
    fp_chars: List[str] = []
    fn_chars: List[str] = []
    tp = fp = fn = 0
    for ch in sorted(keys):
        p = pred_counter.get(ch, 0)
        g = gt_counter.get(ch, 0)
        m = min(p, g)
        if m:
            tp_chars.extend([ch] * m)
            tp += m
        if p > g:
            fp_chars.extend([ch] * (p - g))
            fp += p - g
        if g > p:
            fn_chars.extend([ch] * (g - p))
            fn += g - p
    return ("".join(tp_chars), "".join(fp_chars), "".join(fn_chars), tp, fp, fn)


def micro_f1(tp: int, fp: int, fn: int) -> float:
    denom = (2 * tp) + fp + fn
    if denom == 0:
        return 1.0
    return (2 * tp) / denom


def class_total_payload(tp: int, fp: int, fn: int) -> Dict[str, Any]:
    return {"tp": tp, "fp": fp, "fn": fn, "f1_score": micro_f1(tp, fp, fn)}


def precision(tp: int, fp: int) -> float:
    denom = tp + fp
    if denom == 0:
        return 1.0
    return tp / denom


def recall(tp: int, fn: int) -> float:
    denom = tp + fn
    if denom == 0:
        return 1.0
    return tp / denom


def ensure_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def write_confusion_matrix_png(
    output_path: Path, confusion_matrix: Dict[str, Dict[str, int]]
) -> None:
    gt_labels = set(confusion_matrix.keys())
    pred_labels = set()
    for pred_counts in confusion_matrix.values():
        pred_labels.update(pred_counts.keys())
    axis_labels = sorted(gt_labels | pred_labels)

    # confusion_matrix[gt][pred] -> plot rows as prediction, cols as GT
    matrix = [
        [confusion_matrix.get(gt_label, {}).get(pred_label, 0) for gt_label in axis_labels]
        for pred_label in axis_labels
    ]
    total_count = sum(sum(row) for row in matrix)
    if total_count == 0:
        total_count = 1
    normalized_matrix = [[value / total_count for value in row] for row in matrix]

    ensure_matplotlib_korean_font()

    size = max(6, int(len(axis_labels) * 1.2))
    fig, ax = plt.subplots(figsize=(size, size), constrained_layout=True)
    im = ax.imshow(normalized_matrix, cmap="Blues", vmin=0.0, vmax=1.0)

    ax.set_xticks(range(len(axis_labels)))
    ax.set_yticks(range(len(axis_labels)))
    ax.set_xticklabels(axis_labels, rotation=45, ha="right")
    ax.set_yticklabels(axis_labels)
    ax.set_xlabel("Ground Truth")
    ax.set_ylabel("Prediction")
    ax.set_title("Confusion Matrix (Normalized 0-1)")

    for i, row in enumerate(normalized_matrix):
        for j, value in enumerate(row):
            ax.text(j, i, f"{value:.2f}", ha="center", va="center", color="black")

    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def write_classification_report(
    output_path: Path,
    by_pred_doc_type: Dict[str, List[Tuple[str, str]]],
    confusion_matrix: Dict[str, Dict[str, int]],
) -> None:
    wb = ensure_workbook()
    for doc_type in sorted(by_pred_doc_type.keys()):
        ws_name = doc_type if doc_type else "unknown"
        if len(ws_name) > 31:
            ws_name = ws_name[:31]
        base = ws_name
        idx = 2
        while ws_name in wb.sheetnames:
            suffix = f"_{idx}"
            ws_name = (base[: 31 - len(suffix)] + suffix)[:31]
            idx += 1

        ws = wb.create_sheet(ws_name)
        ws.append(["파일명", "is_correct"])
        for file_name, is_correct in sorted(by_pred_doc_type[doc_type], key=lambda x: x[0]):
            ws.append([file_name, is_correct])

    cm_ws = wb.create_sheet("confusion_matrix")
    gt_labels = set(confusion_matrix.keys())
    pred_labels = set()
    for pred_counts in confusion_matrix.values():
        pred_labels.update(pred_counts.keys())
    axis_labels = sorted(gt_labels | pred_labels)

    cm_ws.append(["gt\\pred"] + axis_labels)
    for gt_label in axis_labels:
        row = [gt_label]
        for pred_label in axis_labels:
            row.append(confusion_matrix.get(gt_label, {}).get(pred_label, 0))
        cm_ws.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_classification_rows(
    pred_dir: Path, gt_dir: Path
) -> Tuple[Dict[str, List[Tuple[str, str]]], Dict[str, Dict[str, int]]]:
    pred_files = xlsx_files_by_stem(pred_dir)
    gt_files = xlsx_files_by_stem(gt_dir)

    gt_doc_type_by_file: Dict[str, str] = {}
    for gt_doc_type, gt_path in gt_files.items():
        gt_kv = read_kv_sheet(gt_path)
        for file_name in gt_kv.keys():
            gt_doc_type_by_file[file_name] = gt_doc_type

    pred_doc_type_by_file: Dict[str, str] = {}
    by_pred_doc_type: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for pred_doc_type, pred_path in pred_files.items():
        pred_kv = read_kv_sheet(pred_path)
        for file_name in pred_kv.keys():
            pred_doc_type_by_file[file_name] = pred_doc_type
            is_correct = "O" if gt_doc_type_by_file.get(file_name) == pred_doc_type else "X"
            by_pred_doc_type[pred_doc_type].append((file_name, is_correct))

    confusion: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    all_files = sorted(set(gt_doc_type_by_file.keys()) | set(pred_doc_type_by_file.keys()))
    for file_name in all_files:
        gt_label = gt_doc_type_by_file.get(file_name, "__missing_gt__")
        pred_label = pred_doc_type_by_file.get(file_name, "__missing_pred__")
        confusion[gt_label][pred_label] += 1

    confusion_plain = {gt: dict(preds) for gt, preds in confusion.items()}
    return by_pred_doc_type, confusion_plain


def write_doc_type_report(
    out_path: Path,
    file_rows: List[
        Tuple[
            str,
            Dict[str, str],
            Dict[str, str],
            Dict[str, Dict[str, List[str]]],
            Dict[str, Dict[str, List[str]]],
        ]
    ],
) -> Tuple[int, int, int]:
    wb = ensure_workbook()

    all_kv_keys = sorted(
        {
            key
            for _, pred_kv, gt_kv, _, _ in file_rows
            for key in list(pred_kv.keys()) + list(gt_kv.keys())
        }
    )
    kv_ws = wb.create_sheet("kv")
    kv_header = ["파일명", "total"]
    for key in all_kv_keys:
        kv_header.extend(
            [f"pred-{key}", f"gt-{key}", f"tp-{key}", f"fp-{key}", f"fn-{key}"]
        )
    kv_ws.append(kv_header)

    all_table_keys = sorted(
        {
            key
            for _, _, _, pred_tables, gt_tables in file_rows
            for key in list(pred_tables.keys()) + list(gt_tables.keys())
        }
    )
    table_class_keys: Dict[str, List[str]] = {}
    for table_key in all_table_keys:
        classes = sorted(
            {
                class_key
                for _, _, _, pred_tables, gt_tables in file_rows
                for class_key in list(pred_tables.get(table_key, {}).keys())
                + list(gt_tables.get(table_key, {}).keys())
            }
        )
        table_class_keys[table_key] = classes

    table_rows_by_key: Dict[str, List[List[Any]]] = {k: [] for k in all_table_keys}
    doc_tp = doc_fp = doc_fn = 0

    for file_name, pred_kv, gt_kv, pred_tables, gt_tables in file_rows:
        file_tp = file_fp = file_fn = 0

        kv_row: List[Any] = [file_name]
        kv_total_tp = kv_total_fp = kv_total_fn = 0
        for key in all_kv_keys:
            pred_text = pred_kv.get(key, "")
            gt_text = gt_kv.get(key, "")
            _, _, _, tp, fp, fn = diff_strings(pred_text, gt_text)
            kv_total_tp += tp
            kv_total_fp += fp
            kv_total_fn += fn
            kv_row.extend([pred_text, gt_text, tp, fp, fn])
            file_tp += tp
            file_fp += fp
            file_fn += fn
        kv_row.insert(
            1,
            json.dumps(
                class_total_payload(kv_total_tp, kv_total_fp, kv_total_fn),
                ensure_ascii=False,
            ),
        )
        kv_ws.append(kv_row)

        for table_key in all_table_keys:
            row: List[Any] = [file_name]
            table_total_tp = table_total_fp = table_total_fn = 0
            for class_key in table_class_keys[table_key]:
                pred_list = pred_tables.get(table_key, {}).get(class_key, [])
                gt_list = gt_tables.get(table_key, {}).get(class_key, [])
                pred_text = "".join(pred_list)
                gt_text = "".join(gt_list)
                _, _, _, tp, fp, fn = diff_strings(pred_text, gt_text)
                table_total_tp += tp
                table_total_fp += fp
                table_total_fn += fn
                file_tp += tp
                file_fp += fp
                file_fn += fn
                row.extend(
                    [
                        json.dumps(pred_list, ensure_ascii=False),
                        json.dumps(gt_list, ensure_ascii=False),
                        tp,
                        fp,
                        fn,
                    ]
                )
            row.insert(
                1,
                json.dumps(
                    class_total_payload(
                        table_total_tp, table_total_fp, table_total_fn
                    ),
                    ensure_ascii=False,
                ),
            )
            table_rows_by_key[table_key].append(row)

        doc_tp += file_tp
        doc_fp += file_fp
        doc_fn += file_fn

    for table_key in all_table_keys:
        ws_name = table_key if table_key else "table"
        if len(ws_name) > 31:
            ws_name = ws_name[:31]
        # De-duplicate sheet names if truncation collides.
        original = ws_name
        idx = 2
        while ws_name in wb.sheetnames:
            suffix = f"_{idx}"
            ws_name = (original[: 31 - len(suffix)] + suffix)[:31]
            idx += 1

        ws = wb.create_sheet(ws_name)
        header = ["파일명", "total"]
        for class_key in table_class_keys[table_key]:
            header.extend(
                [
                    f"pred-{class_key}",
                    f"gt-{class_key}",
                    f"tp-{class_key}",
                    f"fp-{class_key}",
                    f"fn-{class_key}",
                ]
            )
        ws.append(header)
        for row in table_rows_by_key[table_key]:
            ws.append(row)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return doc_tp, doc_fp, doc_fn


def write_summary_workbook(
    output_path: Path, summary_rows: List[Tuple[str, int, int, int]]
) -> None:
    wb = ensure_workbook()
    ws = wb.create_sheet("summary")
    ws.append(["doc_type", "TP", "FP", "FN", "Precision", "Recall", "F1-Score"])
    for doc_type, tp, fp, fn in summary_rows:
        ws.append(
            [doc_type, tp, fp, fn, precision(tp, fp), recall(tp, fn), micro_f1(tp, fp, fn)]
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_tasks(
    pred_dir: Path, gt_dir: Path
) -> Dict[
    str,
    List[
        Tuple[
            str,
            Dict[str, str],
            Dict[str, str],
            Dict[str, Dict[str, List[str]]],
            Dict[str, Dict[str, List[str]]],
        ]
    ],
]:
    if not pred_dir.is_dir():
        raise NotADirectoryError(f"pred directory not found: {pred_dir}")
    if not gt_dir.is_dir():
        raise NotADirectoryError(f"gt directory not found: {gt_dir}")

    pred_files = xlsx_files_by_stem(pred_dir)
    gt_files = xlsx_files_by_stem(gt_dir)
    common = sorted(set(pred_files) & set(gt_files))
    if not common:
        raise ValueError("No matching XLSX filenames between pred and gt directories.")

    by_doc_type: Dict[
        str,
        List[
            Tuple[
                str,
                Dict[str, str],
                Dict[str, str],
                Dict[str, Dict[str, List[str]]],
                Dict[str, Dict[str, List[str]]],
            ]
        ],
    ] = defaultdict(list)

    for stem in common:
        pred_kv_by_file = read_kv_sheet(pred_files[stem])
        gt_kv_by_file = read_kv_sheet(gt_files[stem])
        pred_tables_by_sheet = read_table_sheets(pred_files[stem])
        gt_tables_by_sheet = read_table_sheets(gt_files[stem])

        doc_type = stem
        file_names = sorted(set(pred_kv_by_file) | set(gt_kv_by_file))
        for file_name in file_names:
            pred_kv = pred_kv_by_file.get(file_name, {})
            gt_kv = gt_kv_by_file.get(file_name, {})

            pred_tables: Dict[str, Dict[str, List[str]]] = {}
            gt_tables: Dict[str, Dict[str, List[str]]] = {}
            for sheet_name in sorted(
                set(pred_tables_by_sheet.keys()) | set(gt_tables_by_sheet.keys())
            ):
                pred_tables[sheet_name] = pred_tables_by_sheet.get(sheet_name, {}).get(
                    file_name, {}
                )
                gt_tables[sheet_name] = gt_tables_by_sheet.get(sheet_name, {}).get(
                    file_name, {}
                )

            by_doc_type[doc_type].append(
                (file_name, pred_kv, gt_kv, pred_tables, gt_tables)
            )
    return by_doc_type


def main() -> None:
    args = parse_args()
    grouped = build_tasks(args.pred_dir, args.gt_dir)
    gt_files = xlsx_files_by_stem(args.gt_dir)
    args.eval_dir.mkdir(parents=True, exist_ok=True)
    summary_rows: List[Tuple[str, int, int, int]] = []
    for doc_type, rows in sorted(grouped.items()):
        out_path = args.eval_dir / f"{safe_filename(doc_type)}.xlsx"
        doc_tp, doc_fp, doc_fn = write_doc_type_report(out_path, rows)
        gt_path = gt_files.get(doc_type)
        if gt_path is not None and kv_classes_in_sheet(gt_path):
            summary_rows.append((doc_type, doc_tp, doc_fp, doc_fn))
        print(f"Saved: {out_path}")

    summary_path = args.eval_dir / "summary.xlsx"
    write_summary_workbook(summary_path, summary_rows)
    print(f"Saved: {summary_path}")

    by_pred_doc_type, confusion_matrix = build_classification_rows(args.pred_dir, args.gt_dir)
    classification_path = args.eval_dir / "classification.xlsx"
    write_classification_report(classification_path, by_pred_doc_type, confusion_matrix)
    print(f"Saved: {classification_path}")
    confusion_png_path = args.eval_dir / "confusion_matrix.png"
    write_confusion_matrix_png(confusion_png_path, confusion_matrix)
    print(f"Saved: {confusion_png_path}")


if __name__ == "__main__":
    main()
